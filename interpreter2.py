import numpy as np
from openpyxl import load_workbook,Workbook
from datetime import datetime, timedelta





class Tablet:
    def __init__(self, name: str, box_amount: int, tablets_left: int, daily) -> None:
        self.name = name
        self.box_amount = box_amount
        self.tablets_left = tablets_left
        self.daily = daily

    def __str__(self) -> str:
        return (f"{self.name} with {self.tablets_left} tablets left. {self.daily} to be taken daily. "
                f"Boxes of {self.box_amount}")

# Load the workbook and select the active worksheet
wb = load_workbook('Data.xlsx')
ws = wb.active

# Define the rows to process (corrected to use commas)
rows_to_process = [2,3, 4, 5, 6, 7, 8]

# List to store instances of Tablet
tablets = []



start_date = datetime(2024,8,27)
print(start_date)

# Read data from the worksheet
print("Reading data from 'Data.xlsx':")
for row_index, row in enumerate(ws.iter_rows(min_row=min(rows_to_process), max_row=max(rows_to_process), values_only=True), start=1):
    if row_index in rows_to_process:
        # Extract cell values
        cur_name = row[0] if row[0] is not None else ''
        cur_tablets_left = row[1] if row[1] is not None else 0
        cur_box_amount = row[2] if row[2] is not None else 0
        cur_daily = row[3] if row[3] is not None else 0

        # Create an instance of Tablet
        cur_tablet = Tablet(cur_name, cur_box_amount, cur_tablets_left, cur_daily)
        tablets.append(cur_tablet)

# Print all created instances
for tablet in tablets:
    print(tablet, "\n")




""" If you need to edit the amount taken per day, you need to edit this function"""
def prednisone5_func(target_date):


    # Get the current date
    current_date = datetime.now()
    # Calculate the difference in months
    year_diff = target_date.year - current_date.year
    month_diff = target_date.month - current_date.month

    # Total difference in months
    total_month_diff = year_diff * 12 + month_diff

    if total_month_diff < 0:
        return 0

    # Calculate the score
    score = max(10 - total_month_diff, 0)

    # Use the score to work out how many tablets to take
    
    if score == 10:
        return 2
    elif 5 <= score < 10:
        return 1
    else:
        return 0


def prednisone1_func(target_date):


    # Get the current date
    current_date = datetime.now()
    # Calculate the difference in months
    year_diff = target_date.year - current_date.year
    month_diff = target_date.month - current_date.month

    # Total difference in months
    total_month_diff = year_diff * 12 + month_diff

    if total_month_diff < 0:
        return 0

    # Calculate the score
    score = max(10 - total_month_diff, 0)

    # Use the score to work out how many tablets to take
    
    if score == 10:
        return 0
    elif 5 <= score < 10:
        return score - 5
    else:
        return score
    
def thiamine_func(target_date):
    return 2

def lansoprazole_func(target_date):
    return 1

def furosemide_func(target_date):
    return 1

def metaprolol_func(target_date):
    return 1

# # Test for pred
# date_ = datetime(2023, 12, 20)
# print("mg taken is ",prednisone1_func(date_)+5*prednisone5_func(date_))



def tablet_check(date: datetime, tablets):
    # Date 10 days from now
    check_date = date + timedelta(days=10)

    try:
        wb_current = load_workbook('CurrentData.xlsx')
    except FileNotFoundError:
        wb_current = Workbook()

    
    ws_current = wb_current.active
    ws_current.title = "Tablet Status"
    
    # Header row
    ws_current.append(["Tablet Name", "Tablets Left in 10 days", "Status"])

    
    for tablet in tablets:
        if start_date > check_date:
            # Don't do anything if start_date is after the check date
            continue
        
        # Calculate the projected number of tablets left in 10 days
        projected_tablets_left = tablet.tablets_left
        current_date = start_date
        
        while current_date <= check_date:
            # Dynamically call the function corresponding to the tablet name
            function_name = f"{tablet.name}_func"
            if function_name in globals():
                decrement = globals()[function_name](current_date)
                projected_tablets_left -= decrement


            current_date += timedelta(days=1)
        
        if projected_tablets_left < 0:
            status = "Will run out in 10 days"
        else:
            status = "Safe for the next 10 days"

        ws_current.append([tablet.name, projected_tablets_left, status])
    
    # Save the updated workbook
    wb_current.save('CurrentData.xlsx')


tablet_check(datetime(2024, 9, 27), tablets)

