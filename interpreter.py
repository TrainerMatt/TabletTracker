import numpy as np
from openpyxl import load_workbook
from datetime import datetime, timedelta





class Tablet:
    def __init__(self, name: str, box_amount: int, tablets_left: int, daily) -> None:
        self.name = name
        self.box_amount = box_amount
        self.tablets_left = tablets_left
        self.daily = daily

    def __str__(self) -> str:
        return (f"{self.name} with {self.tablets_left} tablets left. {self.daily} to be taken daily. "
                f"Boxes of {self.box_amount}")

# Load the workbook and select the active worksheet
wb = load_workbook('Data.xlsx')
ws = wb.active

# Define the rows to process (corrected to use commas)
rows_to_process = [2,3, 4, 5, 6, 7, 8]

# List to store instances of Tablet
tablets = []



start_date = datetime(2024,8,27)
print(start_date)

# Read data from the worksheet
print("Reading data from 'Data.xlsx':")
for row_index, row in enumerate(ws.iter_rows(min_row=min(rows_to_process), max_row=max(rows_to_process), values_only=True), start=1):
    if row_index in rows_to_process:
        # Extract cell values
        cur_name = row[0] if row[0] is not None else ''
        cur_tablets_left = row[1] if row[1] is not None else 0
        cur_box_amount = row[2] if row[2] is not None else 0
        cur_daily = row[3] if row[3] is not None else 0

        # Create an instance of Tablet
        cur_tablet = Tablet(cur_name, cur_box_amount, cur_tablets_left, cur_daily)
        tablets.append(cur_tablet)

# Print all created instances
for tablet in tablets:
    print(tablet, "\n")




""" If you need to edit the amount taken per day, you need to edit this function"""
def prednisone5_func(target_date):


    # Get the current date
    current_date = datetime.now()
    # Calculate the difference in months
    year_diff = target_date.year - current_date.year
    month_diff = target_date.month - current_date.month

    # Total difference in months
    total_month_diff = year_diff * 12 + month_diff

    if total_month_diff < 0:
        return 0

    # Calculate the score
    score = max(10 - total_month_diff, 0)

    # Use the score to work out how many tablets to take
    
    if score == 10:
        return 2
    elif 5 <= score < 10:
        return 1
    else:
        return 0


def prednisone1_func(target_date):


    # Get the current date
    current_date = datetime.now()
    # Calculate the difference in months
    year_diff = target_date.year - current_date.year
    month_diff = target_date.month - current_date.month

    # Total difference in months
    total_month_diff = year_diff * 12 + month_diff

    if total_month_diff < 0:
        return 0

    # Calculate the score
    score = max(10 - total_month_diff, 0)

    # Use the score to work out how many tablets to take
    
    if score == 10:
        return 0
    elif 5 <= score < 10:
        return score - 5
    else:
        return score
    
def thiamine_func(target_date):
    return 2

def lansoprazole_func(target_date):
    return 1

def furosemide_func(target_date):
    return 1

def metaprolol_func(target_date):
    return 1

# # Test for pred
# date_ = datetime(2023, 12, 20)
# print("mg taken is ",prednisone1_func(date_)+5*prednisone5_func(date_))




def tablet_check(date: datetime, tablets):
    for tablet in tablets:
        if start_date > date:
            # Don't do anything if start_date is after the given date
            continue
        else:
            # Iterate over each date between start_date and the given date
            current_date = start_date
            while current_date <= date:
                # Dynamically call the function corresponding to the tablet name
                function_name = f"{tablet.name}_func"
                if function_name in globals():
                    print("here")
                    # Assuming the function is defined and available in the global scope
                    decrement = globals()[function_name](current_date)
                    tablet.tablets_left -= decrement
                
                # Move to the next date
                current_date += timedelta(days=1)
            
            # Update the tablet with the new count of tablets left
            print(f"Updated {tablet.name}: {tablet.tablets_left} tablets left")

tablet_check(datetime(2024,12,8),tablets)


